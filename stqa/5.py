import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "students.xlsx")

def create_initial_excel_file():
    if not os.path.exists(EXCEL_PATH):
        print(f"'{os.path.basename(EXCEL_PATH)}' not found. Creating a new file...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Student Grades"
        headers = ["ID", "Student Name", "Grade", "Status"]
        sheet.append(headers)
        for i in range(1, 11):
            sheet.append([i, f"Student {i}", "C", "Pending Update"])
        workbook.save(EXCEL_PATH)
        print("Sample 'students.xlsx' created successfully.")

def main():
    create_initial_excel_file()
    try:
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"Error: Could not find the Excel file at {EXCEL_PATH}")
        return

    print("\nInitializing WebDriver...")
    driver = webdriver.Chrome()
    try:
        driver.get("https://www.w3schools.com/html/html_tables.asp")
        driver.maximize_window()
        time.sleep(2)

        print("Extracting data from W3Schools table...")
        rows = driver.find_elements(By.XPATH, '//*[@id="customers"]/tbody/tr')
        web_data = []
        for row in rows[1:]:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 2:
                company = cells[0].text
                contact = cells[1].text
                web_data.append({'name': company, 'detail': contact})

        student_count = min(10, len(web_data))
        print(f"Found {len(web_data)} records. Will update {student_count} records in Excel.\n")

        for i in range(student_count):
            excel_row_index = i + 2
            data_to_update = web_data[i]
            student_id = sheet.cell(row=excel_row_index, column=1).value
            new_student_name = data_to_update['name']
            new_grade = "A+"
            new_status = f"Updated with contact: {data_to_update['detail']}"
            sheet.cell(row=excel_row_index, column=2, value=new_student_name)
            sheet.cell(row=excel_row_index, column=3, value=new_grade)
            sheet.cell(row=excel_row_index, column=4, value=new_status)
            print(f"Record Updated: ID={student_id}, Name='{new_student_name}', Grade='{new_grade}'")

        workbook.save(EXCEL_PATH)
        print(f"\nSuccessfully saved all updates to '{os.path.basename(EXCEL_PATH)}'.")

        print("\nVerifying updated records in Excel:")
        for i in range(2, student_count + 2):
            student_name = sheet.cell(row=i, column=2).value
            status = sheet.cell(row=i, column=4).value
            print(f"Row {i-1}: Name='{student_name}', Status='{status}'")

    except NoSuchElementException:
        print("Error: Could not find the required elements on the webpage.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    finally:
        driver.quit()
        print("\nData Manipulation and Verification Completed Successfully.")

if __name__ == "__main__":
    main()